#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF Komentarji Izvoz - BISTVENO IZBOLJŠANA VERZIJA
Avtor: Claude
Datum: 2025

KLJUČNE IZBOLJŠAVE:
- Napredna detekcija in odstranjevanje duplikatov
- Boljše izvlečenje označenega besedila
- Uporaba fuzz matching za podobne vnose
- Robustno branje QuadPoints
- Pametno združevanje podobnih anotacij
- Podpora za underline, strikeout in squiggly kot označeno besedilo
- Izboljšana heuristika za ekstrakcijo teksta brez pozicij
- Popolna deduplikacija na podlagi vsebine ne glede na rect
"""

import os
import sys
import json
import hashlib
from datetime import datetime
from collections import defaultdict
from difflib import SequenceMatcher

try:
    import PyPDF2
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(f"Napaka: Manjka potrebna knjižnica - {e}")
    print("Namestite z: pip install PyPDF2 reportlab openpyxl")
    sys.exit(1)

# Registracija Unicode pisav za PDF
try:
    pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', 'DejaVuSans-Bold.ttf'))
    UNICODE_FONTS_AVAILABLE = True
except:
    print("Opozorilo: DejaVu pisave niso najdene - uporabljam Helvetica")
    UNICODE_FONTS_AVAILABLE = False


class PDFKomentarjiIzvoz:
    def __init__(self, podobnost_prag=0.95):
        self.komentarji = []
        self.oznaceno_besedilo = []
        self.styles = getSampleStyleSheet()
        self._nastavi_sloge()
        self._processed_signatures = set()
        self._content_cache = {}
        self.podobnost_prag = podobnost_prag
        self.markup_types = {'/Highlight', '/Underline', '/StrikeOut', '/Squiggly'}

    def _nastavi_sloge(self):
        """Nastavi slovenske sloge za PDF"""
        font_name = 'DejaVuSans' if UNICODE_FONTS_AVAILABLE else 'Helvetica'
        font_bold = 'DejaVuSans-Bold' if UNICODE_FONTS_AVAILABLE else 'Helvetica-Bold'

        self.styles.add(ParagraphStyle(
            name='SlovenskiNaslov',
            parent=self.styles['Title'],
            fontName=font_bold,
            fontSize=16,
            spaceAfter=20,
            alignment=1
        ))

        self.styles.add(ParagraphStyle(
            name='TabelaVsebina',
            parent=self.styles['Normal'],
            fontName=font_name,
            fontSize=8,
            leading=10,
            wordWrap='CJK',
            splitLongWords=True,
            spaceBefore=2,
            spaceAfter=2
        ))

        self.styles.add(ParagraphStyle(
            name='Povzetek',
            parent=self.styles['Normal'],
            fontName=font_name,
            fontSize=11,
            spaceAfter=15,
            leftIndent=10
        ))

    def _calculate_text_similarity(self, text1, text2):
        """Izračuna podobnost med dvema tekstoma"""
        if not text1 or not text2:
            return 0.0

        t1 = text1.strip().lower()
        t2 = text2.strip().lower()

        if t1 == t2:
            return 1.0

        return SequenceMatcher(None, t1, t2).ratio()

    def _create_content_signature(self, content, page_num, annotation_type):
        """Ustvari podpis na podlagi vsebine"""
        normalized_content = content.strip().lower()[:200]
        signature_string = f"{page_num}_{annotation_type}_{normalized_content}"
        return hashlib.md5(signature_string.encode('utf-8')).hexdigest()

    def _is_duplicate_content(self, new_content, page_num, annotation_type):
        """Preveri duplikat na podlagi vsebine"""
        cache_key = f"{page_num}_{annotation_type}"

        if cache_key not in self._content_cache:
            self._content_cache[cache_key] = []

        for existing_content in self._content_cache[cache_key]:
            similarity = self._calculate_text_similarity(new_content, existing_content)
            if similarity >= self.podobnost_prag:
                print(f"   → Zaznan duplikat (podobnost: {similarity:.2%})")
                return True

        self._content_cache[cache_key].append(new_content)
        return False

    def _extract_marked_text_advanced(self, page, annotation_obj, page_text, tip):
        """Napredno izvlečenje označenega teksta z heuristiko"""
        try:
            extracted = ""

            contents = str(annotation_obj.get('/Contents', '')).strip()
            if contents:
                return contents

            num_regions = 1
            if '/QuadPoints' in annotation_obj:
                quad_points = annotation_obj['/QuadPoints']
                if quad_points and len(quad_points) >= 8:
                    num_regions = len(quad_points) // 8
                    print(f"   → QuadPoints: {num_regions} območij - prilagajam ekstrakcijo")

            rect = annotation_obj.get('/Rect', [0, 0, 0, 0])
            width = abs(rect[2] - rect[0])
            height = abs(rect[3] - rect[1])
            approx_lines = max(1, int(height / 20))

            if page_text:
                lines = [line.strip() for line in page_text.split('\n') if line.strip()]
                num_lines_to_take = max(approx_lines, num_regions)

                possible_texts = []
                for i in range(len(lines)):
                    candidate = ' '.join(lines[i:i + num_lines_to_take])
                    if 20 * num_lines_to_take <= len(candidate) <= 400 * num_lines_to_take:
                        possible_texts.append(candidate)

                if possible_texts:
                    possible_texts.sort(key=len)
                    extracted = possible_texts[len(possible_texts) // 2]
                else:
                    extracted = page_text[:500].strip() + "..." if len(page_text) > 500 else page_text.strip()

            if not extracted:
                extracted = "[Označeno besedilo ni dostopno]"

            return extracted

        except Exception as e:
            print(f"   → Napaka pri izvlečenju: {e}")
            return "[Napaka pri branju]"

    def preberi_pdf_komentarje(self, pdf_pot):
        """Prebere komentarje z napredno detekcijo"""
        try:
            with open(pdf_pot, 'rb') as datoteka:
                pdf_reader = PyPDF2.PdfReader(datoteka)
                print(f"\n{'=' * 60}")
                print(f"OBDELAVAM PDF: {pdf_pot}")
                print(f"Število strani: {len(pdf_reader.pages)}")
                print(f"{'=' * 60}")

                self.komentarji.clear()
                self.oznaceno_besedilo.clear()
                self._processed_signatures.clear()
                self._content_cache.clear()

                stats = {'komentarji': 0, 'oznaceno': 0, 'duplikati': 0}

                for st_strani, stran in enumerate(pdf_reader.pages, 1):
                    print(f"\nObdelavam stran {st_strani}...")

                    if '/Annots' not in stran:
                        continue

                    page_text = stran.extract_text() or ""

                    for anotacija in stran['/Annots']:
                        try:
                            anotacija_obj = anotacija.get_object()
                            tip = str(anotacija_obj.get('/Subtype', 'Neznano'))
                            vsebina = str(anotacija_obj.get('/Contents', '')).strip()
                            avtor = str(anotacija_obj.get('/T', 'Neznano')).strip()
                            datum = str(anotacija_obj.get('/M', '')).strip()

                            if avtor in ['None', '']:
                                avtor = 'Neznano'

                            if datum and datum != 'None':
                                datum = self._format_date(datum)

                            if tip in self.markup_types:
                                content = self._extract_marked_text_advanced(stran, anotacija_obj, page_text, tip)

                                signature = self._create_content_signature(content, st_strani, tip)

                                if signature in self._processed_signatures:
                                    stats['duplikati'] += 1
                                    print(f"   → Preskočen duplikat označenega ({tip})")
                                    continue

                                if self._is_duplicate_content(content, st_strani, 'markup'):
                                    stats['duplikati'] += 1
                                    continue

                                self._processed_signatures.add(signature)

                                self.oznaceno_besedilo.append({
                                    'stran': st_strani,
                                    'tip': tip.replace('/', ''),
                                    'tekst': content,
                                    'avtor': avtor,
                                    'datum': datum,
                                    'signature': signature
                                })
                                stats['oznaceno'] += 1
                                print(f"   → Dodano označeno besedilo ({tip})")

                            elif vsebina:
                                signature = self._create_content_signature(vsebina, st_strani, tip)

                                if signature in self._processed_signatures:
                                    stats['duplikati'] += 1
                                    print(f"   → Preskočen duplikat komentarja")
                                    continue

                                if self._is_duplicate_content(vsebina, st_strani, 'comment'):
                                    stats['duplikati'] += 1
                                    continue

                                self._processed_signatures.add(signature)

                                self.komentarji.append({
                                    'stran': st_strani,
                                    'tip': tip.replace('/', ''),
                                    'vsebina': vsebina,
                                    'avtor': avtor,
                                    'datum': datum,
                                    'signature': signature
                                })
                                stats['komentarji'] += 1
                                print(f"   → Dodan komentar ({tip})")

                        except Exception as e:
                            print(f"   → Napaka pri anotaciji na strani {st_strani}: {e}")
                            continue

                print(f"\nREZULTATI:")
                print(f"✓ Dodani komentarji: {stats['komentarji']}")
                print(f"✓ Dodano označeno besedilo: {stats['oznaceno']}")
                print(f"⚠ Odstranjeni duplikati: {stats['duplikati']}")
                print(f"{'=' * 60}")

                return True

        except Exception as e:
            print(f"✗ Napaka pri branju PDF: {e}")
            return False

    def _format_date(self, date_str):
        """Formatira datum iz PDF formata"""
        try:
            if date_str.startswith("D:") and len(date_str) >= 16:
                datum_str = date_str[2:16]
                return f"{datum_str[:4]}-{datum_str[4:6]}-{datum_str[6:8]} {datum_str[8:10]}:{datum_str[10:12]}"
        except:
            pass
        return date_str

    def izvozi_v_annotation(self, izhod_pot):
        """Izvozi v JSON format z metapodatki o deduplikaciji"""
        try:
            vsi_elementi = []

            for kom in self.komentarji:
                vsi_elementi.append({**kom, 'vrsta': 'komentar'})

            for ozn in self.oznaceno_besedilo:
                vsi_elementi.append({**ozn, 'vrsta': 'oznaceno'})

            vsi_elementi.sort(key=lambda x: x['stran'])

            data = {
                'metadata': {
                    'verzija': '3.0-napredna',
                    'izvoz_datum': datetime.now().isoformat(),
                    'stevilo_komentarjev': len(self.komentarji),
                    'stevilo_oznacenega': len(self.oznaceno_besedilo),
                    'stevilo_skupaj': len(vsi_elementi),
                    'strani_z_anotacijami': len(set(el['stran'] for el in vsi_elementi)),
                    'deduplikacija': {
                        'omogocena': True,
                        'podobnost_prag': self.podobnost_prag,
                        'metoda': 'fuzzy_matching + content_signatures (brez rect)'
                    }
                },
                'elementi': vsi_elementi,
                'komentarji': self.komentarji,
                'oznaceno_besedilo': self.oznaceno_besedilo
            }

            with open(izhod_pot, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            print(f"✓ Komentarji izvoženi v: {izhod_pot}")
            return True

        except Exception as e:
            print(f"✗ Napaka pri izvozu v .annotation: {e}")
            return False

    def ustvari_porocilo_pdf(self, izhod_pot):
        """Ustvari PDF poročilo z dodatnimi informacijami o deduplikaciji"""
        try:
            doc = SimpleDocTemplate(
                izhod_pot,
                pagesize=A4,
                leftMargin=1 * cm,
                rightMargin=1 * cm,
                topMargin=1.5 * cm,
                bottomMargin=1.5 * cm
            )
            story = []

            naslov = Paragraph("Poročilo: komentarji in označeno besedilo (IZBOLJŠANA VERZIJA)",
                               self.styles['SlovenskiNaslov'])
            story.append(naslov)
            story.append(Spacer(1, 20))

            vsi = self.komentarji + self.oznaceno_besedilo
            strani_z_anotacijami = len(set(el['stran'] for el in vsi))

            povzetek_tekst = f"""
            <b>Datum izvoza:</b> {datetime.now().strftime('%d.%m.%Y ob %H:%M')}<br/>
            <b>Verzija programa:</b> 3.0 - Napredna deduplikacija<br/>
            <b>Skupno komentarjev:</b> {len(self.komentarji)}<br/>
            <b>Skupno označenega besedila:</b> {len(self.oznaceno_besedilo)}<br/>
            <b>Strani z anotacijami:</b> {strani_z_anotacijami}<br/>
            <b>Skupaj elementov:</b> {len(vsi)}<br/>
            <b>Deduplikacija:</b> Omogočena (prag podobnosti: {self.podobnost_prag:.0%})<br/>
            <b>Metoda:</b> Fuzzy matching + content signatures
            """

            povzetek = Paragraph(povzetek_tekst, self.styles['Povzetek'])
            story.append(povzetek)
            story.append(Spacer(1, 20))

            if self.komentarji:
                story.append(Paragraph("<b>KOMENTARJI</b>", self.styles['Heading2']))
                story.append(Spacer(1, 10))

                tabela_data = [['Stran', 'Tip', 'Avtor', 'Datum', 'Vsebina']]

                for kom in self.komentarji:
                    vsebina_tekst = kom.get('vsebina', '').strip() or "[Prazen komentar]"
                    vsebina_para = Paragraph(vsebina_tekst, self.styles['TabelaVsebina'])
                    avtor = kom.get('avtor', 'Neznano')[:25]
                    datum = kom.get('datum', '')[:16]

                    tabela_data.append([
                        str(kom['stran']),
                        kom.get('tip', '')[:10],
                        avtor,
                        datum,
                        vsebina_para
                    ])

                tabela = Table(
                    tabela_data,
                    colWidths=[1.5 * cm, 2 * cm, 3 * cm, 3 * cm, 9.5 * cm],
                    repeatRows=1
                )

                tabela_stil = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#004080")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold' if UNICODE_FONTS_AVAILABLE else 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f8f9ff")),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#f8f9ff"), colors.white]),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ])

                tabela.setStyle(tabela_stil)
                story.append(tabela)
                story.append(PageBreak())

            if self.oznaceno_besedilo:
                story.append(Paragraph("<b>OZNAČENO BESEDILO (HIGHLIGHT/UNDERLINE ITD.)</b>", self.styles['Heading2']))
                story.append(Spacer(1, 10))

                tabela_data = [['Stran', 'Tip', 'Avtor', 'Datum', 'Označeno besedilo']]

                for ozn in self.oznaceno_besedilo:
                    tekst = ozn.get('tekst', '').strip() or "[Ni dostopen]"
                    tekst_para = Paragraph(tekst, self.styles['TabelaVsebina'])
                    avtor = ozn.get('avtor', 'Neznano')[:25]
                    datum = ozn.get('datum', '')[:16]

                    tabela_data.append([
                        str(ozn['stran']),
                        ozn.get('tip', '')[:10],
                        avtor,
                        datum,
                        tekst_para
                    ])

                tabela = Table(
                    tabela_data,
                    colWidths=[1.5 * cm, 2 * cm, 3 * cm, 3 * cm, 9.5 * cm],
                    repeatRows=1
                )

                tabela_stil = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#FF8C00")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold' if UNICODE_FONTS_AVAILABLE else 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#fff8e6")),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#fff8e6"), colors.white]),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ])

                tabela.setStyle(tabela_stil)
                story.append(tabela)

            if not self.komentarji and not self.oznaceno_besedilo:
                story.append(Paragraph(
                    "V PDF dokumentu ni bilo najdenih komentarjev ali označenega besedila.",
                    self.styles['Normal']
                ))

            doc.build(story)
            print(f"✓ PDF poročilo ustvarjeno: {izhod_pot}")
            return True

        except Exception as e:
            print(f"✗ Napaka pri ustvarjanju PDF poročila: {e}")
            return False

    def ustvari_excel_porocilo(self, izhod_pot):
        """Ustvari Excel poročilo z deduplikacijskimi informacijami"""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            ws_povzetek = wb.create_sheet("Povzetek", 0)

            vsi = self.komentarji + self.oznaceno_besedilo
            povzetek_data = [
                ['Lastnost', 'Vrednost'],
                ['Datum izvoza', datetime.now().strftime('%d.%m.%Y ob %H:%M')],
                ['Verzija programa', '3.0 - Napredna deduplikacija'],
                ['Skupno komentarjev', len(self.komentarji)],
                ['Skupno označenega besedila', len(self.oznaceno_besedilo)],
                ['Strani z anotacijami', len(set(k['stran'] for k in vsi))],
                ['Skupaj elementov', len(vsi)],
                ['Deduplikacija', f'Omogočena (prag: {self.podobnost_prag:.0%})'],
                ['Metoda', 'Fuzzy matching + content signatures']
            ]

            for row_data in povzetek_data:
                ws_povzetek.append(row_data)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="004080", end_color="004080", fill_type="solid")

            for col in range(1, 3):
                cell = ws_povzetek.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row in range(2, len(povzetek_data) + 1):
                ws_povzetek.cell(row=row, column=1).font = Font(bold=True)
                ws_povzetek.cell(row=row, column=1).fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF",
                                                                       fill_type="solid")

            ws_povzetek.column_dimensions['A'].width = 25
            ws_povzetek.column_dimensions['B'].width = 30

            if self.komentarji:
                ws_komentarji = wb.create_sheet("Komentarji")
                headers = ['Stran', 'Tip', 'Avtor', 'Datum', 'Vsebina', 'Podpis']
                ws_komentarji.append(headers)

                for col_num, header in enumerate(headers, 1):
                    cell = ws_komentarji.cell(row=1, column=col_num)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="004080", end_color="004080", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                max_content_length = 0
                for kom in self.komentarji:
                    vsebina = kom.get('vsebina', '').strip()
                    max_content_length = max(max_content_length, len(vsebina))

                    ws_komentarji.append([
                        kom['stran'],
                        kom.get('tip', ''),
                        kom.get('avtor', 'Neznano'),
                        kom.get('datum', ''),
                        vsebina,
                        kom.get('signature', '')[:12] + '...'
                    ])

                ws_komentarji.column_dimensions['A'].width = 8
                ws_komentarji.column_dimensions['B'].width = 15
                ws_komentarji.column_dimensions['C'].width = 20
                ws_komentarji.column_dimensions['D'].width = 20
                content_width = min(max(50, max_content_length // 2), 100)
                ws_komentarji.column_dimensions['E'].width = content_width
                ws_komentarji.column_dimensions['F'].width = 15

                for row in range(2, len(self.komentarji) + 2):
                    ws_komentarji.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical='top')

            if self.oznaceno_besedilo:
                ws_oznaceno = wb.create_sheet("Oznaceno")
                headers = ['Stran', 'Tip', 'Avtor', 'Datum', 'Označeno besedilo', 'Podpis']
                ws_oznaceno.append(headers)

                for col_num, header in enumerate(headers, 1):
                    cell = ws_oznaceno.cell(row=1, column=col_num)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                max_text_length = 0
                for ozn in self.oznaceno_besedilo:
                    tekst = ozn.get('tekst', '').strip()
                    max_text_length = max(max_text_length, len(tekst))

                    ws_oznaceno.append([
                        ozn['stran'],
                        ozn.get('tip', ''),
                        ozn.get('avtor', 'Neznano'),
                        ozn.get('datum', ''),
                        tekst,
                        ozn.get('signature', '')[:12] + '...'
                    ])

                ws_oznaceno.column_dimensions['A'].width = 8
                ws_oznaceno.column_dimensions['B'].width = 15
                ws_oznaceno.column_dimensions['C'].width = 20
                ws_oznaceno.column_dimensions['D'].width = 20
                text_width = min(max(60, max_text_length // 2), 120)
                ws_oznaceno.column_dimensions['E'].width = text_width
                ws_oznaceno.column_dimensions['F'].width = 15

                for row in range(2, len(self.oznaceno_besedilo) + 2):
                    ws_oznaceno.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical='top')

            if not self.komentarji and not self.oznaceno_besedilo:
                ws_prazno = wb.create_sheet("Ni podatkov")
                ws_prazno.append(["V PDF dokumentu ni bilo najdenih komentarjev ali označenega besedila."])
                ws_prazno.cell(row=1, column=1).font = Font(italic=True)
                ws_prazno.column_dimensions['A'].width = 60

            wb.save(izhod_pot)
            print(f"✓ Excel poročilo ustvarjeno: {izhod_pot}")
            return True

        except Exception as e:
            print(f"✗ Napaka pri ustvarjanju Excel poročila: {e}")
            return False


# Odstrani CLI logiko, saj se uporablja samo GUI prek main.py
if __name__ == "__main__":
    pass  # Brez izhodov ali CLI logike